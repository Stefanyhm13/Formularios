import sqlite3
from datetime import datetime
import openpyxl
import shutil
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill
from backend.calculosA import procesar_cuestionario as procesar_cuestionario
from backend.calculosB import procesar_cuestionario_B as procesar_cuestionario_B
from backend.calculosE import procesar_cuestionario_extralaboral as procesar_cuestionario_Extralaboral
from backend.estres import procesar_cuestionario_estres
from backend.rutas import obtener_cuestionarios


FACTORES_TRANSFORMACION = {
    "A + Extralaboral": 616,
    "B + Extralaboral": 512
}

CLASIFICACION_CUESTIONARIOS = {
    "A + Extralaboral": [
        (0.0, 18.8, "Sin riesgo o riesgo despreciable"),
        (18.9, 24.4, "Riesgo bajo"),
        (24.5, 29.5, "Riesgo medio"),
        (29.6, 35.4, "Riesgo alto"),
        (35.5, 100, "Riesgo muy alto")
    ],
    "B + Extralaboral": [
        (0.0, 19.9, "Sin riesgo o riesgo despreciable"),
        (20.0, 24.8, "Riesgo bajo"),
        (24.9, 29.5, "Riesgo medio"),
        (29.6, 35.4, "Riesgo alto"),
        (35.5, 100, "Riesgo muy alto")
    ]
}

def crear_base_datos():
    """Crea las tablas necesarias en la base de datos SQLite en la ruta especificada."""
    # Definir la ruta completa de la base de datos
    db_path = Path(r"C:\Users\practicante.rrhh\Desktop\cuestio_extralab\data\evaluacion_psicosocial.db")

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

     # Tabla para información de los usuarios
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT,
            identificacion TEXT UNIQUE,
            area TEXT
        )
    ''')

    # Tabla para información básica de la evaluación
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS evaluaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha DATETIME,
            tipo_empleado TEXT,
            usuario_id INTEGER,
            FOREIGN KEY (usuario_id) REFERENCES usuarios (id)
        )
    ''')

    # Tabla para resultados de cuestionarios
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS resultados_cuestionarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            evaluacion_id INTEGER,
            usuario_id INTEGER,
            tipo_cuestionario TEXT,  -- 'A', 'B', 'Extralaboral', 'Estres', 'A + Extralaboral', 'B + Extralaboral'
            forma_cuestionario TEXT, -- 'A' o 'B' para los cuestionarios base
            puntaje_bruto REAL,
            puntaje_transformado REAL,
            clasificacion TEXT,
            FOREIGN KEY (evaluacion_id) REFERENCES evaluaciones (id),
            FOREIGN KEY (usuario_id) REFERENCES usuarios (id)
        )
    ''')

    # Tabla para detalles de dominios
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS dominios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            resultado_id INTEGER,
            usuario_id INTEGER,
            dominio TEXT,
            puntaje_bruto REAL,
            puntaje_transformado REAL,
            clasificacion TEXT,
            FOREIGN KEY (resultado_id) REFERENCES resultados_cuestionarios (id),
            FOREIGN KEY (usuario_id) REFERENCES usuarios (id)
        )
    ''')

    # Tabla para detalles de dimensiones
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS dimensiones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            resultado_id INTEGER,
            usuario_id INTEGER,
            dominio TEXT,
            dimension TEXT,
            puntaje_bruto REAL,
            puntaje_transformado REAL,
            clasificacion TEXT,
            FOREIGN KEY (resultado_id) REFERENCES resultados_cuestionarios (id),
            FOREIGN KEY (usuario_id) REFERENCES usuarios (id)
        )
    ''')

    conn.commit()
    conn.close()

def guardar_en_db(tipo_empleado, nombre_empleado, cedula, area, datos_a, datos_b, datos_extralaboral, datos_estres):
    """Guarda todos los resultados en la base de datos existente."""
    # Ruta fija a la base de datos en la carpeta 'data'
    db_path = Path(r"C:\Users\practicante.rrhh\Desktop\cuestio_extralab\data\evaluacion_psicosocial.db")

    # Conectar a la base de datos
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    fecha_actual = datetime.now()

    # Insertar información del usuario
    cursor.execute('''
        INSERT INTO usuarios (nombre, Identificacion, area)
        VALUES (?, ?, ?)
    ''', (nombre_empleado, cedula, area))
    usuario_id = cursor.lastrowid

    # Insertar información básica de la evaluación
    cursor.execute('''
        INSERT INTO evaluaciones (fecha, tipo_empleado, usuario_id)
        VALUES (?, ?, ?)
    ''', (fecha_actual, tipo_empleado, usuario_id))
    evaluacion_id = cursor.lastrowid

    # Función auxiliar para insertar resultado de cuestionario
    def insertar_resultado_cuestionario(tipo_cuestionario, forma_cuestionario, puntaje_bruto, 
                                        puntaje_transformado, clasificacion):
        cursor.execute('''
            INSERT INTO resultados_cuestionarios 
            (evaluacion_id, usuario_id, tipo_cuestionario, forma_cuestionario, 
             puntaje_bruto, puntaje_transformado, clasificacion)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (evaluacion_id, usuario_id, tipo_cuestionario, forma_cuestionario,
              puntaje_bruto, puntaje_transformado, clasificacion))
        return cursor.lastrowid

    # Función auxiliar para guardar dominios y dimensiones
    def guardar_dominios_dimensiones(datos, resultado_id):
        puntajes, puntaje_total, puntajes_transformados, clasificaciones, _, _ = datos
        for dominio, dimensiones in puntajes.items():
            if isinstance(dimensiones, dict):
                # Es un dominio con dimensiones
                cursor.execute('''
                    INSERT INTO dominios 
                    (resultado_id, usuario_id, dominio, puntaje_bruto, puntaje_transformado, clasificacion)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (resultado_id, usuario_id, dominio, sum(bruto for bruto in dimensiones.values() if bruto != "Sin puntaje válido"),
                      puntajes_transformados[dominio]["TOTAL_DOMINIO"],
                      clasificaciones[dominio]["TOTAL_DOMINIO"]))
                dominio_id = cursor.lastrowid

                # Guardar dimensiones
                for dimension, bruto in dimensiones.items():
                    if dimension != "TOTAL_DOMINIO":
                        cursor.execute('''
                            INSERT INTO dimensiones 
                            (resultado_id, usuario_id, dominio, dimension, puntaje_bruto, puntaje_transformado, clasificacion)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', (resultado_id, usuario_id, dominio, dimension, bruto,
                              puntajes_transformados[dominio][dimension],
                              clasificaciones[dominio][dimension]))
            else:
                # Es un dominio sin dimensiones
                cursor.execute('''
                    INSERT INTO dominios 
                    (resultado_id, usuario_id, dominio, puntaje_bruto, puntaje_transformado, clasificacion)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (resultado_id, usuario_id, dominio, dimensiones,
                      puntajes_transformados[dominio],
                      clasificaciones[dominio]))

    # Guardar resultados según el tipo de empleado
    if tipo_empleado.upper() == 'A':
        # Para jefes, guardar cuestionario A
        if datos_a:
            resultado_id_a = insertar_resultado_cuestionario(
                'Intralaboral', 'A',
                datos_a[1],  # puntaje_total
                datos_a[4],  # transformado_total
                datos_a[5]   # clasificacion_total
            )
            guardar_dominios_dimensiones(datos_a, resultado_id_a)
        
        # Marcar cuestionario B como no aplicable
        insertar_resultado_cuestionario(
            'Intralaboral', 'B',
            None,
            None,
            'NO APLICA'
        )
    else:
        # Para otros empleados, guardar cuestionario B
        if datos_b:
            resultado_id_b = insertar_resultado_cuestionario(
                'Intralaboral', 'B',
                datos_b[1],
                datos_b[4],
                datos_b[5]
            )
            guardar_dominios_dimensiones(datos_b, resultado_id_b)
        
        # Marcar cuestionario A como no aplicable
        insertar_resultado_cuestionario(
            'Intralaboral', 'A',
            None,
            None,
            'NO APLICA'
        )

    # Guardar resultados extralaborales
    if datos_extralaboral:
        resultado_id_extra = insertar_resultado_cuestionario(
            'Extralaboral', None,
            datos_extralaboral[1],
            datos_extralaboral[4],
            datos_extralaboral[5]
        )
        guardar_dominios_dimensiones(datos_extralaboral, resultado_id_extra)

    # Guardar resultados de estrés
    if datos_estres:
        insertar_resultado_cuestionario(
            'Estrés', None,
            datos_estres[0],
            datos_estres[1],
            datos_estres[2]
        )

    # Guardar resultados combinados
    if tipo_empleado.lower() == 'a' or 'A':
        # Para jefes, guardar A + Extralaboral
        puntaje_A_Extralaboral = datos_a[4] + datos_extralaboral[4]
        transformado_A_Extralaboral = round((puntaje_A_Extralaboral / FACTORES_TRANSFORMACION["A + Extralaboral"]) * 100, 1)
        clasificacion_A_Extralaboral = clasificar_puntaje(transformado_A_Extralaboral, 
                                                         CLASIFICACION_CUESTIONARIOS["A + Extralaboral"])
        
        insertar_resultado_cuestionario(
            'A + Extralaboral', 'A',
            puntaje_A_Extralaboral,
            transformado_A_Extralaboral,
            clasificacion_A_Extralaboral
        )
        
        # Marcar B + Extralaboral como no aplicable
        insertar_resultado_cuestionario(
            'B + Extralaboral', 'B',
            None,
            None,
            'NO APLICA'
        )
    else:
        # Para otros empleados, guardar B + Extralaboral
        puntaje_B_Extralaboral = datos_b[4] + datos_extralaboral[4]
        transformado_B_Extralaboral = round((puntaje_B_Extralaboral / FACTORES_TRANSFORMACION["B + Extralaboral"]) * 100, 1)
        clasificacion_B_Extralaboral = clasificar_puntaje(transformado_B_Extralaboral, 
                                                         CLASIFICACION_CUESTIONARIOS["B + Extralaboral"])
        
        insertar_resultado_cuestionario(
            'B + Extralaboral', 'B',
            puntaje_B_Extralaboral,
            transformado_B_Extralaboral,
            clasificacion_B_Extralaboral
        )
        
        # Marcar A + Extralaboral como no aplicable
        insertar_resultado_cuestionario(
            'A + Extralaboral', 'A',
            None,
            None,
            'NO APLICA'
        )

    conn.commit()
    conn.close()

    return "Datos guardados exitosamente en la base de datos."

# Las funciones existentes se mantienen igual
def obtener_color_clasificacion(clasificacion):
    """
    Devuelve el color hexadecimal basado en el nivel de clasificación.
    """
    colores = {
        "Sin riesgo o riesgo despreciable": "FF27AE60",
        "Riesgo bajo": "48FF68",
        "Riesgo medio": "FFFF00",
        "Riesgo alto": "FF9900",
        "Riesgo muy alto": "FF0000",
        "Muy bajo": "FF27AE60",
        "Bajo": "48FF68",
        "Medio": "FFFF00",
        "Alto": "FF9900",
        "Muy alto": "FF0000",
    }
    return colores.get(clasificacion, "FFFFFF")

def clasificar_puntaje(puntaje, clasificacion_rangos):
    for rango_min, rango_max, nivel in clasificacion_rangos:
        if rango_min <= puntaje <= rango_max:
            return nivel
    return "Clasificación no encontrada"

# Función principal para calcular y clasificar puntajes combinados
def calcular_puntaje_total(transformado_total, transformado_total_Extralaboral,tipo_empleado):
    
    if tipo_empleado.lower() == 'a' or 'A':
        puntaje_A_Extralaboral = round(transformado_total+ transformado_total_Extralaboral)
        transformado_A_Extralaboral = round((puntaje_A_Extralaboral / FACTORES_TRANSFORMACION["A + Extralaboral"]) * 100, 1)
        clasificacion_A_Extralaboral = clasificar_puntaje(transformado_A_Extralaboral, CLASIFICACION_CUESTIONARIOS["A + Extralaboral"])

        return list((puntaje_A_Extralaboral, transformado_A_Extralaboral, clasificacion_A_Extralaboral))

    elif tipo_empleado.lower() == 'b' or 'B':
    
        puntaje_B_Extralaboral = round(transformado_total+ transformado_total_Extralaboral)
        transformado_B_Extralaboral = round((puntaje_B_Extralaboral / FACTORES_TRANSFORMACION["B + Extralaboral"]) * 100, 1)
        clasificacion_B_Extralaboral = clasificar_puntaje(transformado_B_Extralaboral, CLASIFICACION_CUESTIONARIOS["B + Extralaboral"])

        return list((puntaje_B_Extralaboral, transformado_B_Extralaboral, clasificacion_B_Extralaboral))

    
# Función para ajustar el ancho de las columnas automáticamente
def ajustar_columnas(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Obtener la letra de la columna
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
        
def generar_excel(puntaje_bruto_estres, puntaje_transformado_estres, clasificacion_estres, identificacion, tipo_empleado,cuestionarios,respuestas_totales,respuestas_a,respuestas_b,respuestas_extralaboral):
    wb = openpyxl.Workbook()

    ws_total = wb.create_sheet("Total General")
    escribir_encabezado(ws_total, ["Cuestionarios evaluados", "Bruto", "Transformado", "Clasificación"], "00A9DF")

    if tipo_empleado.lower() == 'a' or 'A':
        ws_a = wb.active
        ws_a.title = "Cuestionario A"
        datos_a = respuestas_a
        escribir_datos_cuestionario(ws_a, datos_a, "A", ws_total)

        ws_extralaboral = wb.create_sheet("Cuestionario Extralaboral")
        datos_extralaboral = respuestas_extralaboral
        escribir_datos_cuestionario(ws_extralaboral, datos_extralaboral, "Extralaboral", ws_total)

        pdf_rutas = [
            cuestionarios[0],
            cuestionarios[2],
            cuestionarios[3]
        ]

    else:
        ws_b = wb.active
        ws_b.title = "Cuestionario B"
        datos_b = respuestas_b
        escribir_datos_cuestionario(ws_b, datos_b, "B", ws_total)

        ws_extralaboral = wb.create_sheet("Cuestionario Extralaboral")
        datos_extralaboral =respuestas_extralaboral
        escribir_datos_cuestionario(ws_extralaboral, datos_extralaboral, "Extralaboral", ws_total)

        pdf_rutas = [
            cuestionarios[1],
            cuestionarios[2],
            cuestionarios[3]
        ]

    if tipo_empleado.lower() == 'a' or 'A':
        ws_total.append(["A + Extralaboral", respuestas_totales[0], respuestas_totales[1],respuestas_totales[2]])
        color = obtener_color_clasificacion(respuestas_totales[2])
        ws_total[f"D{ws_total.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    else:
        ws_total.append(["B + Extralaboral", respuestas_totales[0], respuestas_totales[1], respuestas_totales[2]])
        color = obtener_color_clasificacion(respuestas_totales[2])
        ws_total[f"D{ws_total.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    ws_total.append(["Estrés", puntaje_bruto_estres, puntaje_transformado_estres, clasificacion_estres])
    color = obtener_color_clasificacion(clasificacion_estres)
    ws_total[f"D{ws_total.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    ajustar_columnas(ws_total)

    documentos_path = Path.home() / "Documents"
    carpeta_cuestionarios = documentos_path / "Cuestionarios"
    carpeta_identificacion = carpeta_cuestionarios / identificacion
    carpeta_identificacion.mkdir(parents=True, exist_ok=True)

    archivo_total_path = carpeta_identificacion / "Resultados_Cuestionarios.xlsx"
    wb.save(archivo_total_path)
    print(f"Archivo Excel generado: {archivo_total_path}")

    for pdf in pdf_rutas:
        pdf_path = Path(pdf)
        if pdf_path.exists():
            shutil.copy(pdf_path, carpeta_identificacion)
            print(f"Archivo PDF copiado: {pdf_path}")

    mns="Datos guardados exitosamente en el archivo Excel."
    return mns

def escribir_encabezado(ws, columnas, fill_color="00A9DF"):
    ws.append(columnas)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

def aplicar_estilo_fila(ws, color_hex, bold=False):
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        if bold:
            cell.font = Font(bold=True)

def escribir_datos_cuestionario(ws, datos, cuestionario,ws_total):
    encabezado_fill = PatternFill(start_color="00A9DF", end_color="00A9DF", fill_type="solid")
    dominio_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    dimension_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    puntajes, puntaje_total, puntajes_transformados, clasificaciones, transformado_total, clasificacion_total = datos

    ws.append(["Cuestionario", cuestionario])
    ws.append(["Dominio/Dimensión", "Bruto", "Transformado", "Clasificación"])
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = encabezado_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if isinstance(next(iter(puntajes.values())), dict):
        for dominio, dimensiones in puntajes.items():
            total_bruto_dominio = sum(dimensiones.values())
            total_transformado_dominio = puntajes_transformados[dominio]["TOTAL_DOMINIO"]
            clasificacion_dominio = clasificaciones[dominio]["TOTAL_DOMINIO"]

            ws.append([f"Dominio: {dominio}", total_bruto_dominio, total_transformado_dominio, clasificacion_dominio])
            for cell in ws[ws.max_row]:
                cell.fill = dominio_fill
                cell.font = Font(bold=True)

            color = obtener_color_clasificacion(clasificacion_dominio)
            ws[f"D{ws.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            for dimension, bruto in dimensiones.items():
                if dimension != "TOTAL_DOMINIO":
                    transformado = puntajes_transformados[dominio][dimension]
                    clasificacion = clasificaciones[dominio][dimension]
                    ws.append([f"  Dimensión: {dimension}", bruto, transformado, clasificacion])
                    for cell in ws[ws.max_row]:
                        cell.fill = dimension_fill

                    color = obtener_color_clasificacion(clasificacion)
                    ws[f"D{ws.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    else:
        for dimension, bruto in puntajes.items():
            if bruto is not None:
                transformado = puntajes_transformados[dimension]
                clasificacion = clasificaciones[dimension]
                ws.append([dimension, bruto, transformado, clasificacion])
                for cell in ws[ws.max_row]:
                    cell.fill = dimension_fill

                color = obtener_color_clasificacion(clasificacion)
                ws[f"D{ws.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    ws.append(["TOTAL", puntaje_total, transformado_total, clasificacion_total])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FF000000")
        cell.fill = encabezado_fill

    color = obtener_color_clasificacion(clasificacion_total)
    ws[f"D{ws.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    ajustar_columnas(ws)
    
   

