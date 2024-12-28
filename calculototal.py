# Importar transformados totales de los archivos específicos
from calculosA import transformado_total as transformado_total_A
from calculosB import transformado_total as transformado_total_B
from calculosE import transformado_total as transformado_total_Extralaboral 

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from calculosA import procesar_cuestionario as procesar_cuestionario_A
from calculosB import procesar_cuestionario_B as procesar_cuestionario_B
from calculosE import procesar_cuestionario_extralaboral as procesar_cuestionario_Extralaboral
from estres import procesar_cuestionario_estres


# Factores de transformación para los cuestionarios combinados
FACTORES_TRANSFORMACION = {
    "A + Extralaboral": 616,
    "B + Extralaboral": 512
}

# Clasificación de riesgo para los cuestionarios combinados
CLASIFICACION_CUESTIONARIOS = {
    "A + Extralaboral": [
        (0.0, 18.8, "Sin riesgo o riesgo despreciable"),
        (18.9, 24.4, "Riesgo bajo"),
        (24.5, 29.5, "Riesgo medio"),
        (29.6, 35.4, "Riesgo alto"),
        (35.5, 100, "Riesgo muy alto")
    ],
    "B + Extralaboral": [
        (0.0, 19.9, "Sin riesgo o riesgo despreciable"),
        (20.0, 24.8, "Riesgo bajo"),
        (24.9, 29.5, "Riesgo medio"),
        (29.6, 35.4, "Riesgo alto"),
        (35.5, 100, "Riesgo muy alto")
    ]
}

def obtener_color_clasificacion(clasificacion):
    """
    Devuelve el color hexadecimal basado en el nivel de clasificación.
    """
    colores = {
        # Clasificaciones generales
        "Sin riesgo o riesgo despreciable": "FF27AE60",  # Verde intenso
        "Riesgo bajo": "48FF68",  # Verde menos encendido
        "Riesgo medio": "FFFF00",  # Amarillo
        "Riesgo alto": "FF9900",  # Naranja
        "Riesgo muy alto": "FF0000",  # Rojo
        
        # Clasificaciones específicas del cuestionario de estrés
        "Muy bajo": "FF27AE60",  # Verde intenso
        "Bajo": "48FF68",  # Verde menos encendido
        "Medio": "FFFF00",  # Amarillo
        "Alto": "FF9900",  # Naranja
        "Muy alto": "FF0000",  # Rojo
    }
    return colores.get(clasificacion, "FFFFFF")  # Por defecto blanco si no coincide



# Función para clasificar un puntaje según rangos
def clasificar_puntaje(puntaje, clasificacion_rangos):
    for rango_min, rango_max, nivel in clasificacion_rangos:
        if rango_min <= puntaje <= rango_max:
            return nivel
    return "Clasificación no encontrada"

# Función principal para calcular y clasificar puntajes combinados
def calcular_puntaje_total():
    # Calcular combinaciones de Forma A + Extralaboral y Forma B + Extralaboral
    puntaje_A_Extralaboral = transformado_total_A + transformado_total_Extralaboral
    puntaje_B_Extralaboral = transformado_total_B + transformado_total_Extralaboral

    # Transformar los puntajes
    transformado_A_Extralaboral = round((puntaje_A_Extralaboral / FACTORES_TRANSFORMACION["A + Extralaboral"]) * 100, 1)
    transformado_B_Extralaboral = round((puntaje_B_Extralaboral / FACTORES_TRANSFORMACION["B + Extralaboral"]) * 100, 1)

    # Clasificar los puntajes transformados
    clasificacion_A_Extralaboral = clasificar_puntaje(transformado_A_Extralaboral, CLASIFICACION_CUESTIONARIOS["A + Extralaboral"])
    clasificacion_B_Extralaboral = clasificar_puntaje(transformado_B_Extralaboral, CLASIFICACION_CUESTIONARIOS["B + Extralaboral"])

    return (puntaje_A_Extralaboral, transformado_A_Extralaboral, clasificacion_A_Extralaboral,
            puntaje_B_Extralaboral, transformado_B_Extralaboral, clasificacion_B_Extralaboral)

# Función para ajustar el ancho de las columnas automáticamente
def ajustar_columnas(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Obtener la letra de la columna
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

# Función para generar el archivo Excel
def generar_excel(puntaje_bruto_estres, puntaje_transformado_estres, clasificacion_estres):
    wb = openpyxl.Workbook()

    # Hoja para Cuestionario A
    ws_a = wb.active
    ws_a.title = "Cuestionario A"
    datos_a = procesar_cuestionario_A()
    escribir_datos_cuestionario(ws_a, datos_a, "A")

    # Hoja para Cuestionario B
    ws_b = wb.create_sheet("Cuestionario B")
    datos_b = procesar_cuestionario_B()
    escribir_datos_cuestionario(ws_b, datos_b, "B")

    # Hoja para Cuestionario Extralaboral
    ws_extralaboral = wb.create_sheet("Cuestionario Extralaboral")
    datos_extralaboral = procesar_cuestionario_Extralaboral()
    escribir_datos_cuestionario(ws_extralaboral, datos_extralaboral, "Extralaboral")

    # Hoja para el Total General
    ws_total = wb.create_sheet("Total General")
    escribir_datos_totales(ws_total, puntaje_bruto_estres, puntaje_transformado_estres, clasificacion_estres)

    # Guardar el archivo
    wb.save("Resultados_Cuestionarios.xlsx")
    print("Archivo Excel generado: Resultados_Cuestionarios.xlsx")


from openpyxl.styles import Font, Alignment, PatternFill

# Función para escribir un encabezado formateado en una hoja
def escribir_encabezado(ws, columnas, fill_color="00A9DF"):
    """Escribe un encabezado formateado en la hoja de Excel."""
    ws.append(columnas)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Función para aplicar un estilo a una fila
def aplicar_estilo_fila(ws, color_hex, bold=False):
    """Aplica un estilo específico a la última fila escrita."""
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        if bold:
            cell.font = Font(bold=True)

def escribir_datos_cuestionario(ws, datos, cuestionario):
    encabezado_fill = PatternFill(start_color="00A9DF", end_color="00A9DF", fill_type="solid")
    dominio_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    dimension_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    puntajes, puntaje_total, puntajes_transformados, clasificaciones, transformado_total, clasificacion_total = datos

    # Escribir encabezado
    ws.append(["Cuestionario", cuestionario])
    ws.append(["Dominio/Dimensión", "Bruto", "Transformado", "Clasificación"])
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = encabezado_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Si el cuestionario tiene dominios
    if isinstance(next(iter(puntajes.values())), dict):
        for dominio, dimensiones in puntajes.items():
            total_bruto_dominio = sum(dimensiones.values())
            total_transformado_dominio = puntajes_transformados[dominio]["TOTAL_DOMINIO"]
            clasificacion_dominio = clasificaciones[dominio]["TOTAL_DOMINIO"]

            # Escribir los resultados del dominio
            ws.append([f"Dominio: {dominio}", total_bruto_dominio, total_transformado_dominio, clasificacion_dominio])
            for cell in ws[ws.max_row]:
                cell.fill = dominio_fill
                cell.font = Font(bold=True)

            # Aplicar color a la celda de clasificación
            color = obtener_color_clasificacion(clasificacion_dominio)
            ws[f"D{ws.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Escribir las dimensiones dentro del dominio
            for dimension, bruto in dimensiones.items():
                if dimension != "TOTAL_DOMINIO":
                    transformado = puntajes_transformados[dominio][dimension]
                    clasificacion = clasificaciones[dominio][dimension]
                    ws.append([f"  Dimensión: {dimension}", bruto, transformado, clasificacion])
                    for cell in ws[ws.max_row]:
                        cell.fill = dimension_fill

                    # Aplicar color a la celda de clasificación
                    color = obtener_color_clasificacion(clasificacion)
                    ws[f"D{ws.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    else:
        for dimension, bruto in puntajes.items():
            if bruto is not None:
                transformado = puntajes_transformados[dimension]
                clasificacion = clasificaciones[dimension]
                ws.append([dimension, bruto, transformado, clasificacion])
                for cell in ws[ws.max_row]:
                    cell.fill = dimension_fill

                # Aplicar color a la celda de clasificación
                color = obtener_color_clasificacion(clasificacion)
                ws[f"D{ws.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Total del cuestionario
    ws.append(["TOTAL", puntaje_total, transformado_total, clasificacion_total])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FF000000")
        cell.fill = encabezado_fill

    # Aplicar color a la celda de clasificación total
    color = obtener_color_clasificacion(clasificacion_total)
    ws[f"D{ws.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Ajustar ancho de columnas
    ajustar_columnas(ws)


def escribir_datos_totales(ws, puntaje_bruto_estres, puntaje_transformado_estres, clasificacion_estres):
    """Escribe los datos totales de los cuestionarios combinados."""
    encabezado_fill = "00A9DF"

    # Cálculos existentes
    puntaje_A_Extralaboral = transformado_total_A + transformado_total_Extralaboral
    puntaje_B_Extralaboral = transformado_total_B + transformado_total_Extralaboral

    transformado_A_Extralaboral = round((puntaje_A_Extralaboral / FACTORES_TRANSFORMACION["A + Extralaboral"]) * 100, 1)
    transformado_B_Extralaboral = round((puntaje_B_Extralaboral / FACTORES_TRANSFORMACION["B + Extralaboral"]) * 100, 1)

    clasificacion_A_Extralaboral = clasificar_puntaje(transformado_A_Extralaboral, CLASIFICACION_CUESTIONARIOS["A + Extralaboral"])
    clasificacion_B_Extralaboral = clasificar_puntaje(transformado_B_Extralaboral, CLASIFICACION_CUESTIONARIOS["B + Extralaboral"])

    # Encabezado
    escribir_encabezado(ws, ["Cuestionarios evaluados", "Bruto", "Transformado", "Clasificación"], encabezado_fill)

    # Datos
    datos_totales = [
        ("A + Extralaboral", puntaje_A_Extralaboral, transformado_A_Extralaboral, clasificacion_A_Extralaboral),
        ("B + Extralaboral", puntaje_B_Extralaboral, transformado_B_Extralaboral, clasificacion_B_Extralaboral),
        ("Estrés", puntaje_bruto_estres, puntaje_transformado_estres, clasificacion_estres),
    ]

    for cuestionario, bruto, transformado, clasificacion in datos_totales:
        ws.append([cuestionario, bruto, transformado, clasificacion])

        # Aplicar color a la celda de clasificación
        color = obtener_color_clasificacion(clasificacion)
        ws[f"D{ws.max_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Ajustar ancho de columnas
    ajustar_columnas(ws)



# Ejecutar el script principal
if __name__ == "__main__":
    tipo_empleado = input("Ingrese el tipo de empleado (Jefes / Operarios): ")
    puntaje_bruto_estres, puntaje_transformado_estres, clasificacion_estres = procesar_cuestionario_estres(tipo_empleado)

    if puntaje_bruto_estres is None:
        print("Error: No se pudo procesar el cuestionario de estrés.")
    else:
        generar_excel(puntaje_bruto_estres, puntaje_transformado_estres, clasificacion_estres)
